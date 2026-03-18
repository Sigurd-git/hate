from __future__ import annotations

import logging
from dataclasses import dataclass
from glob import glob
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple

import pandas as pd
from openpyxl import load_workbook


LOGGER = logging.getLogger(__name__)

PROJECT_ROOT = Path(__file__).resolve().parents[1]

MODEL_LABEL_OVERRIDES = {
    "openai_gpt-5.1": "chatgpt5.1",
    "anthropic_claude-opus-4.5": "claude4.5",
}

PARADIGM_FOLDERS = {
    "zeroshot": "zero_shot",
    "cot": "chain_of_thought",
}


@dataclass(frozen=True)
class DatasetSpec:
    language: str
    raw_dataset_path: Path
    outputs_root: Path
    combined_output_path: Path
    f1_output_path: Path
    brier_output_path: Path
    label_column: str = "label/2classes"
    text_column: str = "text"


DATASET_SPECS: Dict[str, DatasetSpec] = {
    "zh": DatasetSpec(
        language="zh",
        raw_dataset_path=PROJECT_ROOT / "data/5_new_chinesehatedata_2400_balanced.xlsx",
        outputs_root=PROJECT_ROOT / "outputs/natural/5_new_chinesehatedata_2400_balanced/zh",
        combined_output_path=PROJECT_ROOT / "data/8_new_chinesehatedata+zeroshot+cot.xlsx",
        f1_output_path=PROJECT_ROOT / "data/9_new_chinesehatedata_model_f1_scores.xlsx",
        brier_output_path=PROJECT_ROOT / "data/9_new_chinesehatedata_model_brier_scores.xlsx",
    ),
    "en": DatasetSpec(
        language="en",
        raw_dataset_path=PROJECT_ROOT / "data/5_new_englishhatedata_2400_balanced.xlsx",
        outputs_root=PROJECT_ROOT / "outputs/natural/5_new_englishhatedata_2400_balanced/en",
        combined_output_path=PROJECT_ROOT / "data/8_new_englishhatedata_2400+zeroshot+cot.xlsx",
        f1_output_path=PROJECT_ROOT / "data/9_new_englishhatedata_model_f1_scores.xlsx",
        brier_output_path=PROJECT_ROOT / "data/9_new_englishhatedata_model_brier_scores.xlsx",
    ),
}


def normalize_model_label(model_name: str) -> str:
    """Return a stable display label used in metric output columns."""
    return MODEL_LABEL_OVERRIDES.get(model_name, model_name)


def list_excel_files(folder: Path) -> List[Path]:
    """Return all workbook files under a folder in stable order."""
    return sorted(Path(path) for path in glob(str(folder / "*.xlsx")))


def parse_model_and_sample_index(file_path: Path) -> Tuple[str, int]:
    """Parse the model name and numeric sample index from a workbook filename."""
    stem = file_path.stem
    if "_sample_" not in stem:
        raise ValueError(f"Unexpected filename format: {file_path.name}")
    model_name, sample_marker = stem.rsplit("_sample_", 1)
    if not sample_marker.isdigit():
        raise ValueError(f"Unexpected sample marker in filename: {file_path.name}")
    return model_name, int(sample_marker)


def normalize_text_value(text_value: object) -> str:
    """Normalize text values so workbook rows can be matched back to the source dataset."""
    if pd.isna(text_value):
        return ""
    return str(text_value).replace("\r\n", "\n").strip()


def build_text_index(raw_text_series: pd.Series) -> Dict[str, List[int]]:
    """Map each source text string to the list of matching raw row indices."""
    text_index: Dict[str, List[int]] = {}
    for raw_index, raw_text in enumerate(raw_text_series.tolist()):
        normalized_text = normalize_text_value(raw_text)
        text_index.setdefault(normalized_text, []).append(raw_index)
    return text_index


def read_sample_output(file_path: Path) -> Tuple[str, Optional[float], int]:
    """Read the text, score, and sample index from a one-row workbook."""
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
        value_row = next(worksheet.iter_rows(min_row=2, max_row=2, values_only=True))
        row_mapping = dict(zip(header_row, value_row))
    finally:
        workbook.close()

    if "text" not in row_mapping or "sample_index" not in row_mapping:
        raise ValueError(f"Workbook missing required columns: {file_path}")

    normalized_text = normalize_text_value(row_mapping["text"])
    score_value = row_mapping.get("score")
    parsed_sample_index = int(row_mapping["sample_index"])
    parsed_score = None if pd.isna(score_value) else float(score_value)
    return normalized_text, parsed_score, parsed_sample_index


def resolve_raw_index(
    sample_text: str,
    reported_sample_index: int,
    raw_texts: Sequence[str],
    text_to_indices: Dict[str, List[int]],
    assigned_raw_indices: set[int],
    previous_aligned_index: Optional[int],
) -> Tuple[Optional[int], str]:
    """Resolve a workbook row to a raw dataset row using sample index first, then text fallback."""
    if (
        0 <= reported_sample_index < len(raw_texts)
        and raw_texts[reported_sample_index] == sample_text
        and reported_sample_index not in assigned_raw_indices
    ):
        return reported_sample_index, "sample_index"

    candidate_indices = [
        candidate_index
        for candidate_index in text_to_indices.get(sample_text, [])
        if candidate_index not in assigned_raw_indices
    ]
    if not candidate_indices:
        return None, "unresolved"

    if len(candidate_indices) == 1:
        return candidate_indices[0], "text"

    if reported_sample_index in candidate_indices:
        return reported_sample_index, "text_exact_match"

    if previous_aligned_index is not None:
        monotonic_candidates = [
            candidate_index
            for candidate_index in candidate_indices
            if candidate_index >= previous_aligned_index
        ]
        if monotonic_candidates:
            candidate_indices = monotonic_candidates

    chosen_index = min(
        candidate_indices,
        key=lambda candidate_index: abs(candidate_index - reported_sample_index),
    )
    return chosen_index, "text_ambiguous"


def collect_model_scores(
    model_files: Sequence[Path],
    raw_texts: Sequence[str],
    text_to_indices: Dict[str, List[int]],
) -> Tuple[pd.Series, dict]:
    """Collect one model/setting score vector aligned to the raw dataset rows."""
    aligned_scores = pd.Series(pd.NA, index=range(len(raw_texts)), dtype="Float64")
    assigned_raw_indices: set[int] = set()
    previous_aligned_index: Optional[int] = None

    alignment_counts = {
        "sample_index": 0,
        "text": 0,
        "text_exact_match": 0,
        "text_ambiguous": 0,
        "unresolved": 0,
    }

    for file_path in sorted(model_files):
        sample_text, score_value, reported_sample_index = read_sample_output(file_path)
        aligned_index, resolution_method = resolve_raw_index(
            sample_text=sample_text,
            reported_sample_index=reported_sample_index,
            raw_texts=raw_texts,
            text_to_indices=text_to_indices,
            assigned_raw_indices=assigned_raw_indices,
            previous_aligned_index=previous_aligned_index,
        )
        alignment_counts[resolution_method] += 1
        if aligned_index is None:
            continue

        aligned_scores.iloc[aligned_index] = score_value
        assigned_raw_indices.add(aligned_index)
        previous_aligned_index = aligned_index

    alignment_summary = {
        "files": len(model_files),
        "assigned": int(aligned_scores.notna().sum()),
        "missing_rows": int(aligned_scores.isna().sum()),
        **alignment_counts,
    }
    return aligned_scores, alignment_summary


def build_language_evaluation_frame(
    dataset_spec: DatasetSpec,
    save_combined_output: bool = True,
) -> pd.DataFrame:
    """Build an evaluation table with labels plus aligned score columns for every model."""
    LOGGER.info("Loading raw dataset for language=%s from %s", dataset_spec.language, dataset_spec.raw_dataset_path)
    raw_frame = pd.read_excel(dataset_spec.raw_dataset_path)
    evaluation_frame = raw_frame.copy()
    evaluation_frame["sample_index"] = range(len(evaluation_frame))

    raw_texts = [
        normalize_text_value(text_value)
        for text_value in evaluation_frame[dataset_spec.text_column].tolist()
    ]
    text_to_indices = build_text_index(evaluation_frame[dataset_spec.text_column])

    score_columns: List[str] = []
    for paradigm_label in ("zeroshot", "cot"):
        paradigm_folder = dataset_spec.outputs_root / PARADIGM_FOLDERS[paradigm_label]
        paradigm_files = list_excel_files(paradigm_folder)
        model_groups: Dict[str, List[Path]] = {}
        for file_path in paradigm_files:
            model_name, _ = parse_model_and_sample_index(file_path)
            model_groups.setdefault(model_name, []).append(file_path)

        for model_name in sorted(model_groups, key=normalize_model_label):
            output_column = f"{normalize_model_label(model_name)}_{paradigm_label}_score"
            aligned_scores, alignment_summary = collect_model_scores(
                model_files=model_groups[model_name],
                raw_texts=raw_texts,
                text_to_indices=text_to_indices,
            )
            evaluation_frame[output_column] = aligned_scores
            score_columns.append(output_column)
            LOGGER.info(
                "Aligned language=%s model=%s setting=%s files=%s assigned=%s missing_rows=%s "
                "sample_index=%s text=%s text_exact_match=%s text_ambiguous=%s unresolved=%s",
                dataset_spec.language,
                normalize_model_label(model_name),
                paradigm_label,
                alignment_summary["files"],
                alignment_summary["assigned"],
                alignment_summary["missing_rows"],
                alignment_summary["sample_index"],
                alignment_summary["text"],
                alignment_summary["text_exact_match"],
                alignment_summary["text_ambiguous"],
                alignment_summary["unresolved"],
            )

    ordered_columns = [
        column_name
        for column_name in evaluation_frame.columns
        if column_name not in score_columns
    ] + score_columns
    evaluation_frame = evaluation_frame.loc[:, ordered_columns]

    if save_combined_output:
        dataset_spec.combined_output_path.parent.mkdir(parents=True, exist_ok=True)
        evaluation_frame.to_excel(dataset_spec.combined_output_path, index=False)
        LOGGER.info(
            "Saved aligned evaluation table for language=%s to %s",
            dataset_spec.language,
            dataset_spec.combined_output_path,
        )

    return evaluation_frame
